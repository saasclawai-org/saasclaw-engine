"""Import PA tax codes from DCED Excel (.xlsx) or CSV file.

Usage:
    python manage.py import_pa_tax_codes <file.xlsx|file.csv> --year 2026 [--replace]

The DCED Excel format (EitWithCollector_Dyn_Excel) has these columns.
The --replace flag deletes all existing records for the year before importing.
"""
import csv
import sys
from decimal import Decimal

from django.core.management.base import BaseCommand


def clean_psd(val):
    if not val:
        return ''
    return str(val).strip().replace('.', '').replace(' ', '')


def parse_pct(val):
    """Parse rate values. DCED stores as percentages (1.00 = 1%, 0.5 = 0.5%).
    Model stores as decimal (1% = 0.01, 0.5% = 0.005).
    Always divide by 100 because DCED values are always percentages."""
    if val is None or str(val).strip() == '':
        return Decimal('0')
    s = str(val).strip().replace('%', '').replace(',', '').strip()
    if s in ('', '-', 'N/A', 'n/a'):
        return Decimal('0')
    try:
        num = Decimal(s)
    except Exception:
        return Decimal('0')
    return num / Decimal('100')


def parse_dollar(val):
    if val is None or str(val).strip() == '':
        return Decimal('0')
    s = str(val).strip().replace('$', '').replace(',', '').strip()
    if s in ('', '-', 'N/A', 'n/a'):
        return Decimal('0')
    try:
        return Decimal(s)
    except Exception:
        return Decimal('0')

def parse_date(val):
    """Parse date strings in MM/DD/YYYY or YYYY-MM-DD format."""
    if val is None or str(val).strip() == '':
        return None
    from datetime import datetime
    s = str(val).strip()
    if s in ('', '-', 'N/A', 'n/a'):
        return None
    # Try MM/DD/YYYY
    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# (model_field, is_pct, is_dollar)
COLUMN_MAP = {
    'psd code': ('psd_code', False, False),
    'psd_code': ('psd_code', False, False),
    'psd': ('psd_code', False, False),
    'tax collection district name': ('tax_collection_district', False, False),
    'tax collection district': ('tax_collection_district', False, False),
    'tcd': ('tax_collection_district', False, False),
    'county': ('county', False, False),
    'municipality id': ('municipality_id', False, False),
    'municipality': ('municipality', False, False),
    'school district id': ('school_district_id', False, False),
    'school district name': ('school_district', False, False),
    'school district': ('school_district', False, False),
    'resident eit rate': ('municipal_resident_eit_rate', True, False),
    'municipal resident eit rate': ('municipal_resident_eit_rate', True, False),
    'municipal resident eit (%)': ('municipal_resident_eit_rate', True, False),
    'resident eit': ('municipal_resident_eit_rate', True, False),
    'resident eit (%)': ('municipal_resident_eit_rate', True, False),
    'nonresident eit rate': ('municipal_nonresident_eit_rate', True, False),
    'municipal nonresident eit rate': ('municipal_nonresident_eit_rate', True, False),
    'municipal nonresident eit (%)': ('municipal_nonresident_eit_rate', True, False),
    'nonresident eit': ('municipal_nonresident_eit_rate', True, False),
    'nonresident eit (%)': ('municipal_nonresident_eit_rate', True, False),
    'school district eit rate': ('school_district_eit_rate', True, False),
    'school district eit (%)': ('school_district_eit_rate', True, False),
    'school district eit': ('school_district_eit_rate', True, False),
    'school district pit rate': ('school_district_pit_rate', True, False),
    'school district pit (%)': ('school_district_pit_rate', True, False),
    'school district pit': ('school_district_pit_rate', True, False),
    'total resident eit rate': ('total_resident_eit_rate', True, False),
    'total resident income tax (%)': ('total_resident_eit_rate', True, False),
    'total resident eit': ('total_resident_eit_rate', True, False),
    'total resident eit (%)': ('total_resident_eit_rate', True, False),
    'total resident income tax': ('total_resident_eit_rate', True, False),
    'municipal eit lie': ('municipal_eit_lie', False, True),
    'municipal lie amount': ('municipal_eit_lie', False, True),
    'municipal eit lie amount': ('municipal_eit_lie', False, True),
    'municipal lie': ('municipal_eit_lie', False, True),
    'school district eit lie': ('school_district_eit_lie', False, True),
    'school district lie amount': ('school_district_eit_lie', False, True),
    'school district eit lie amount': ('school_district_eit_lie', False, True),
    'school district lie': ('school_district_eit_lie', False, True),
    'municipal lst': ('municipal_lst', False, True),
    'municipal lst ($)': ('municipal_lst', False, True),
    'municipal lst amount': ('municipal_lst', False, True),
    'school district lst': ('school_district_lst', False, True),
    'school district lst ($)': ('school_district_lst', False, True),
    'school district lst amount': ('school_district_lst', False, True),
    'total lst': ('total_lst', False, True),
    'total lst ($)': ('total_lst', False, True),
    'total lst amount': ('total_lst', False, True),
    'municipal lst lie': ('municipal_lst_lie', False, True),
    'municipal lst lie amount': ('municipal_lst_lie', False, True),
    'school district lst lie': ('school_district_lst_lie', False, True),
    'school district lst lie amount': ('school_district_lst_lie', False, True),
    'eit collector': ('eit_collector', False, False),
    'eit collector name': ('eit_collector', False, False),
    'collector': ('eit_collector', False, False),
    'collector address': ('eit_collector_address1', False, False),
    'collector city': ('eit_collector_city', False, False),
    'collector state': ('eit_collector_state', False, False),
    'collector zip': ('eit_collector_zip', False, False),
    'collector phone': ('eit_collector_phone', False, False),
    'collector email': ('eit_collector_email', False, False),
    'collector website': ('eit_collector_website', False, False),
    'date last updated': ('date_last_updated', False, False, True),
    # DCED format (after stripping (percent)/(dollars) suffixes)
    'municipal resident eit': ('municipal_resident_eit_rate', True, False),
    'municipal nonresident eit': ('municipal_nonresident_eit_rate', True, False),
    'municipal nonresident eit effective date': ('municipal_nonresident_eit_effective_date', False, False, True),
    'municipal resident eit effective date': ('municipal_resident_eit_effective_date', False, False, True),
    'school district eit effective date': ('school_district_eit_effective_date', False, False, True),
    'municipal lst effective date': ('municipal_lst_effective_date', False, False, True),
    'school district lst effective date': ('school_district_lst_effective_date', False, False, True),
    'reporting year': ('year', False, False),
    'eit collector address1': ('eit_collector_address1', False, False),
    'eit collector address2': ('eit_collector_address2', False, False),
    'eit collector address3': ('eit_collector_address3', False, False),
    'eit collector city': ('eit_collector_city', False, False),
    'eit collector state': ('eit_collector_state', False, False),
    'eit collector zip': ('eit_collector_zip', False, False),
    'eit collector zip extension': ('eit_collector_zip_extension', False, False),
    'eit collector phone': ('eit_collector_phone', False, False),
    'eit collector phone extension': ('eit_collector_phone_extension', False, False),
    'eit collector toll free phone': ('eit_collector_toll_free_phone', False, False),
    'eit collector fax': ('eit_collector_fax', False, False),
    'eit collector web site': ('eit_collector_website', False, False),
    'municipal lst collector': ('municipal_lst_collector', False, False),
    'school district lst collector': ('school_district_lst_collector', False, False),
}



class Command(BaseCommand):
    help = 'Import PA tax codes from DCED Excel (.xlsx) or CSV file'

    def add_arguments(self, parser):
        parser.add_argument('file', type=str, help='Path to Excel (.xlsx) or CSV file')
        parser.add_argument('--year', type=int, required=True, help='Tax year (e.g., 2026)')
        parser.add_argument('--replace', action='store_true', help='Delete existing records for the year before importing')

    def handle(self, *args, **options):
        from saasclaw_engine.tax_data.models_pa import PaTaxCode

        filepath = options['file']
        year = options['year']
        replace = options['replace']

        if filepath.endswith('.xlsx'):
            rows = self._read_xlsx(filepath)
        elif filepath.endswith('.csv'):
            rows = self._read_csv(filepath)
        else:
            self.stderr.write(self.style.ERROR(f'Unsupported file format: {filepath}'))
            sys.exit(1)

        if not rows or len(rows) < 2:
            self.stderr.write(self.style.ERROR('No data rows found'))
            sys.exit(1)

        import re
        header = [str(c).strip() if c else '' for c in rows[0]]
        # Normalize headers: strip (percent)/(dollars) suffixes for matching
        normalized_header = [re.sub(r'\s*\((?:percent|dollars)\)\s*$', '', h, flags=re.IGNORECASE).strip().lower() for h in header]
        col_map = {}
        unmapped = []
        for i, col in enumerate(normalized_header):
            if col in COLUMN_MAP:
                col_map[i] = COLUMN_MAP[col]
            else:
                unmapped.append(header[i])

        if unmapped:
            self.stdout.write(f'Unmapped columns ({len(unmapped)}): {unmapped[:15]}')

        mapped_fields = set(v[0] for v in col_map.values())
        self.stdout.write(f'Mapped {len(col_map)} columns: {sorted(mapped_fields)}')

        if replace:
            deleted = PaTaxCode.objects.filter(year=year).delete()[0]
            self.stdout.write(f'Deleted {deleted} existing records for year {year}')

        created = 0
        updated = 0
        skipped = 0

        # Get model field names for validation
        model_fields = {f.name for f in PaTaxCode._meta.get_fields()}

        for row in rows[1:]:
            record = {}
            for i, val in enumerate(row):
                if i in col_map:
                    col_info = col_map[i]
                    field = col_info[0]
                    is_pct = col_info[1]
                    is_dollar = col_info[2]
                    is_date = col_info[3] if len(col_info) > 3 else False
                    if field in record:
                        continue
                    # Skip fields not in the model
                    if field not in model_fields:
                        continue
                    if is_pct:
                        record[field] = parse_pct(val)
                    elif is_dollar:
                        record[field] = parse_dollar(val)
                    elif is_date or field.endswith('_date'):
                        record[field] = parse_date(val)
                    else:
                        record[field] = str(val).strip() if val else ''

            psd_code = record.get('psd_code', '').strip()
            if not psd_code:
                skipped += 1
                continue

            psd_code = clean_psd(psd_code)
            record['psd_code'] = psd_code
            record['year'] = year

            
            record.setdefault('tax_collection_district', '')
            record.setdefault('county', '')
            record.setdefault('municipality_id', '')
            record.setdefault('municipality', '')
            record.setdefault('school_district_id', '')
            record.setdefault('school_district', '')
            record.setdefault('municipal_nonresident_eit_rate', Decimal('0'))
            record.setdefault('municipal_resident_eit_rate', Decimal('0'))
            record.setdefault('school_district_eit_rate', Decimal('0'))
            record.setdefault('school_district_pit_rate', Decimal('0'))
            record.setdefault('total_resident_eit_rate', Decimal('0'))
            record.setdefault('municipal_eit_lie', Decimal('0'))
            record.setdefault('school_district_eit_lie', Decimal('0'))
            record.setdefault('municipal_lst', Decimal('0'))
            record.setdefault('school_district_lst', Decimal('0'))
            record.setdefault('total_lst', Decimal('0'))
            record.setdefault('municipal_lst_lie', Decimal('0'))
            record.setdefault('school_district_lst_lie', Decimal('0'))
            record.setdefault('eit_collector', '')
            record.setdefault('eit_collector_address1', '')
            record.setdefault('eit_collector_city', '')
            record.setdefault('eit_collector_state', '')
            record.setdefault('eit_collector_zip', '')
            record.setdefault('eit_collector_phone', '')
            record.setdefault('eit_collector_email', '')
            record.setdefault('eit_collector_website', '')
            # date fields are already parsed by parse_date()

            obj, was_created = PaTaxCode.objects.update_or_create(
                year=year, psd_code=psd_code,
                defaults=record,
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'Imported {created + updated} PA tax codes for year {year} '
            f'({created} created, {updated} updated, {skipped} skipped)'
        ))

    def _read_xlsx(self, filepath):
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()

        # DCED Excel has a junk instruction row 1 and headers in row 2
        # Detect: if row 1 doesn't look like headers, skip it
        if rows and rows[0] and rows[0][0] and 'please scroll' in str(rows[0][0]).lower():
            self.stdout.write('Detected DCED format: skipping instruction row 1')
            rows = rows[1:]  # Now rows[0] is the real header row

        # Normalize header row: strip whitespace
        if rows:
            rows[0] = [str(c).strip() if c else '' for c in rows[0]]
            self.stdout.write(f'Detected columns: {rows[0][:10]}...')

        return rows

    def _read_csv(self, filepath):
        rows = []
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
        return rows